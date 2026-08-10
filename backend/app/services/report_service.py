from datetime import datetime, timezone

from io import BytesIO

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
)
from openpyxl.utils import get_column_letter

from app.models.enums import BookingStatus
from app.repositories.report_repository import (
    ReportRepository,
)
from app.schemas.report import (
    BookingHistoryResponse,
)


class ReportService:

    def __init__(
        self,
        report_repository: ReportRepository,
    ):
        self.report_repository = (
            report_repository
        )

    def _ensure_utc(
        self,
        value: datetime,
    ) -> datetime:

        if value.tzinfo is None:
            return value.replace(
                tzinfo=timezone.utc,
            )

        return value.astimezone(
            timezone.utc,
        )

    def get_booking_history(
        self,
        user_id: int | None = None,
        meeting_room_id: int | None = None,
        status: BookingStatus | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> list[BookingHistoryResponse]:

        if start_date is not None:
            start_date = self._ensure_utc(
                start_date,
            )

        if end_date is not None:
            end_date = self._ensure_utc(
                end_date,
            )

        if (
            start_date is not None
            and end_date is not None
            and start_date > end_date
        ):
            raise ValueError(
                "start_date must be before end_date."
            )

        bookings = (
            self.report_repository
            .get_booking_history(
                user_id=user_id,
                meeting_room_id=meeting_room_id,
                status=status,
                start_date=start_date,
                end_date=end_date,
            )
        )

        return [
            BookingHistoryResponse.model_validate(
                booking,
            )
            for booking in bookings
        ]
        
    def export_booking_history_excel(
        self,
        user_id: int | None = None,
        meeting_room_id: int | None = None,
        status: BookingStatus | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> BytesIO:

        bookings = self.get_booking_history(
            user_id=user_id,
            meeting_room_id=meeting_room_id,
            status=status,
            start_date=start_date,
            end_date=end_date,
        )

        workbook = Workbook()

        worksheet = workbook.active

        if worksheet is None:
            raise RuntimeError(
                "Unable to create Excel worksheet."
            )

        worksheet.title = "Booking History"

        headers = [
            "Booking ID",
            "User ID",
            "Meeting Room ID",
            "Title",
            "Description",
            "Start Time",
            "End Time",
            "Status",
            "Recurrence ID",
            "Created At",
            "Updated At",
        ]

        worksheet.append(headers)

        for booking in bookings:
            worksheet.append(
                [
                    booking.id,
                    booking.user_id,
                    booking.meeting_room_id,
                    booking.title,
                    booking.description,
                    str(booking.start_time),
                    str(booking.end_time),
                    booking.status.value,
                    booking.recurrence_id,
                    str(booking.created_at),
                    str(booking.updated_at),
                ]
            )

        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):

            max_length = 0

            column_letter = get_column_letter(
                column_index
            )

            for row in range(
                1,
                worksheet.max_row + 1,
            ):

                cell = worksheet.cell(
                    row=row,
                    column=column_index,
                )

                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                40,
            )

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        return output
    
    def export_booking_history_pdf(
        self,
        user_id: int | None = None,
        meeting_room_id: int | None = None,
        status: BookingStatus | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> BytesIO:

        bookings = self.get_booking_history(
            user_id=user_id,
            meeting_room_id=meeting_room_id,
            status=status,
            start_date=start_date,
            end_date=end_date,
        )

        output = BytesIO()

        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
        )

        data = [
            [
                "ID",
                "User ID",
                "Room ID",
                "Title",
                "Start Time",
                "End Time",
                "Status",
            ]
        ]

        for booking in bookings:

            data.append(
                [
                    str(booking.id),
                    str(booking.user_id),
                    str(booking.meeting_room_id),
                    booking.title,
                    str(booking.start_time),
                    str(booking.end_time),
                    booking.status.value,
                ]
            )

        table = Table(data)

        table.setStyle(
        TableStyle(
            (
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.grey,
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.black,
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER",
                ),
            )
        )
    )

        document.build(
            [table]
        )

        output.seek(0)

        return output